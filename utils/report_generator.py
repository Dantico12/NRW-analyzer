import pandas as pd
import numpy as np
import os
import uuid
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
import base64
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.marker import DataPoint
import logging

logger = logging.getLogger(__name__)

class ReportGenerator:
    def __init__(self):
        self.reports_dir = "reports"
        os.makedirs(self.reports_dir, exist_ok=True)
        
    def generate_excel_report(self, region_data: list, task_data: list, service_data: list, filename: str) -> str:
        """
        Generate comprehensive Excel report with multiple sheets and charts
        """
        try:
            # Create unique filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_filename = f"analysis_report_{timestamp}.xlsx"
            report_path = os.path.join(self.reports_dir, report_filename)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create summary sheet
            self._create_summary_sheet(wb, region_data, task_data, service_data, filename)
            
            # Create detailed sheets
            self._create_region_sheet(wb, region_data)
            self._create_task_sheet(wb, task_data)
            self._create_service_sheet(wb, service_data)
            
            # Save workbook
            wb.save(report_path)
            
            logger.info(f"Report generated successfully: {report_path}")
            return report_path
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            raise Exception(f"Failed to generate report: {str(e)}")
    
    def _create_summary_sheet(self, wb: Workbook, region_data: list, task_data: list, service_data: list, filename: str):
        """Create summary sheet with overview"""
        ws = wb.create_sheet("Summary", 0)
        
        # Header styling
        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        ws['A1'] = "COMPLAINT ANALYSIS REPORT"
        ws['A1'].font = Font(bold=True, size=18)
        ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        ws['A1'].font = Font(bold=True, size=18, color="FFFFFF")
        ws.merge_cells('A1:E1')
        
        # Report info
        ws['A3'] = "Report Generated:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A4'] = "Source File:"
        ws['B4'] = filename
        
        # Summary statistics
        ws['A6'] = "SUMMARY STATISTICS"
        ws['A6'].font = header_font
        ws['A6'].fill = header_fill
        ws['A6'].font = Font(bold=True, size=14, color="FFFFFF")
        ws.merge_cells('A6:B6')
        
        ws['A8'] = "Total Regions:"
        ws['B8'] = len(region_data)
        ws['A9'] = "Total Task Types:"
        ws['B9'] = len(task_data)
        ws['A10'] = "Total Tasks:"
        ws['B10'] = sum(item['total_tasks'] for item in region_data) if region_data else 0
        
        # Top regions
        ws['A12'] = "TOP 5 REGIONS BY TASK COUNT"
        ws['A12'].font = header_font
        ws['A12'].fill = header_fill
        ws['A12'].font = Font(bold=True, size=14, color="FFFFFF")
        ws.merge_cells('A12:B12')
        
        top_regions = sorted(region_data, key=lambda x: x['total_tasks'], reverse=True)[:5]
        for i, region in enumerate(top_regions, start=14):
            ws[f'A{i}'] = region['REGION']
            ws[f'B{i}'] = region['total_tasks']
        
        # Top tasks
        ws['D12'] = "TOP 5 TASK TYPES"
        ws['D12'].font = header_font
        ws['D12'].fill = header_fill
        ws['D12'].font = Font(bold=True, size=14, color="FFFFFF")
        ws.merge_cells('D12:E12')
        
        top_tasks = sorted(task_data, key=lambda x: x['Count'], reverse=True)[:5]
        for i, task in enumerate(top_tasks, start=14):
            ws[f'D{i}'] = task['Task']
            ws[f'E{i}'] = task['Count']
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_region_sheet(self, wb: Workbook, region_data: list):
        """Create detailed region analysis sheet"""
        ws = wb.create_sheet("Region Analysis")
        
        # Headers
        headers = ["Region", "Total Tasks", "Percentage"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        total_tasks = sum(item['total_tasks'] for item in region_data)
        for row, item in enumerate(region_data, start=2):
            ws.cell(row=row, column=1, value=item['REGION'])
            ws.cell(row=row, column=2, value=item['total_tasks'])
            percentage = (item['total_tasks'] / total_tasks * 100) if total_tasks > 0 else 0
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
        
        # Create chart
        if len(region_data) > 0:
            chart = BarChart()
            chart.title = "Tasks by Region"
            chart.x_axis.title = "Region"
            chart.y_axis.title = "Number of Tasks"
            
            data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(region_data)+1)
            categories_ref = Reference(ws, min_col=1, min_row=2, max_row=len(region_data)+1)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(categories_ref)
            
            ws.add_chart(chart, "E2")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_task_sheet(self, wb: Workbook, task_data: list):
        """Create detailed task analysis sheet"""
        ws = wb.create_sheet("Task Analysis")
        
        # Headers
        headers = ["Task Type", "Count", "Percentage"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        total_count = sum(item['Count'] for item in task_data)
        for row, item in enumerate(task_data, start=2):
            ws.cell(row=row, column=1, value=item['Task'])
            ws.cell(row=row, column=2, value=item['Count'])
            percentage = (item['Count'] / total_count * 100) if total_count > 0 else 0
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
        
        # Create chart
        if len(task_data) > 0:
            chart = BarChart()
            chart.title = "Task Distribution"
            chart.x_axis.title = "Task Type"
            chart.y_axis.title = "Count"
            
            # Limit to top 10 for readability
            chart_data = task_data[:10]
            data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(chart_data)+1)
            categories_ref = Reference(ws, min_col=1, min_row=2, max_row=len(chart_data)+1)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(categories_ref)
            
            ws.add_chart(chart, "E2")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_service_sheet(self, wb: Workbook, service_data: list):
        """Create service analysis sheet"""
        ws = wb.create_sheet("Service Analysis")
        
        # Headers
        headers = ["Region", "Service Type", "Count"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        row = 2
        for region_data in service_data:
            region = region_data['region']
            for service in region_data['services']:
                ws.cell(row=row, column=1, value=region)
                ws.cell(row=row, column=2, value=service['service_type'])
                ws.cell(row=row, column=3, value=service['count'])
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

# Legacy function for backward compatibility
def generate_excel_report(region_data: list, task_data: list, filename: str) -> str:
    """Legacy function wrapper"""
    generator = ReportGenerator()
    return generator.generate_excel_report(region_data, task_data, [], filename)